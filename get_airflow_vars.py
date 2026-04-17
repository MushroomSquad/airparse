#!/usr/bin/env -S uv run
# /// script
# requires-python = ">=3.11"
# dependencies = [
#   "requests",
#   "openpyxl",
#   "docker",
# ]
# ///

"""
Airflow Variable/Connection/Secret extractor
Supports 3 modes:
  --mode api     — remote via Airflow REST API
  --mode local   — local system (reads airflow DB or CLI)
  --mode docker  — local Docker container
"""

import argparse
import json
import subprocess
import sys
from dataclasses import dataclass, field
from pathlib import Path

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


@dataclass
class AirflowData:
    connections: list[dict] = field(default_factory=list)
    variables: list[dict] = field(default_factory=list)
    secrets: list[dict] = field(default_factory=list)


def extract_json(text: str) -> dict | list:
    """Extract first valid JSON object/array from text mixed with logs."""
    # Find first { or [
    start_obj = text.find("{")
    start_arr = text.find("[")
    
    if start_obj == -1 and start_arr == -1:
        raise ValueError("No JSON found in output")
    
    if start_obj == -1:
        start = start_arr
    elif start_arr == -1:
        start = start_obj
    else:
        start = min(start_obj, start_arr)
    
    # Use raw_decode to parse only the first JSON structure
    decoder = json.JSONDecoder()
    obj, _ = decoder.raw_decode(text[start:])
    return obj


# ─────────────────────────── API MODE ────────────────────────────────────────


def fetch_api(base_url: str, user: str, password: str, cookie: str | None = None) -> AirflowData:
    session = requests.Session()
    session.headers["Content-Type"] = "application/json"
    base = base_url.rstrip("/")

    # ── Auth: try all methods in order ──────────────────────────────────────
    authed = False

    # 0. Pre-supplied browser cookie (highest priority — bypasses all login flows)
    if cookie:
        for pair in cookie.split(";"):
            pair = pair.strip()
            if "=" in pair:
                name, _, value = pair.partition("=")
                session.cookies.set(name.strip(), value.strip())
        print(f"[API] Auth: using supplied cookie ({cookie[:40]}…)")
        authed = True

    # 1. Airflow 3.x token endpoint
    if not authed:
        for url, payload in [
            (f"{base}/auth/token",            {"username": user, "password": password}),
            (f"{base}/api/v1/security/login", {"username": user, "password": password, "refresh": False}),
        ]:
            try:
                r = session.post(url, json=payload, timeout=10)
                if r.ok:
                    token = r.json().get("access_token") or r.json().get("jwt_token")
                    if token:
                        session.headers["Authorization"] = f"Bearer {token}"
                        print(f"[API] Auth: JWT via {url}")
                        authed = True
                        break
            except Exception:
                pass

    # 2. Web-form login → session cookie (works when UI uses LDAP/OAuth but API uses session backend)
    if not authed:
        try:
            login_page = session.get(f"{base}/login", timeout=10)
            # Extract CSRF token from login form
            import re
            csrf_match = re.search(r'name="csrf_token"\s+value="([^"]+)"', login_page.text)
            csrf = csrf_match.group(1) if csrf_match else ""
            login_r = session.post(f"{base}/login", data={
                "username":   user,
                "password":   password,
                "csrf_token": csrf,
            }, allow_redirects=True, timeout=10)
            # Check if we landed on the dashboard (not back on /login)
            if login_r.ok and "/login" not in login_r.url:
                print(f"[API] Auth: web-form session cookie (landed on {login_r.url})")
                authed = True
            else:
                print(f"[API] Auth: web-form login redirected back to login page — wrong credentials or SSO")
        except Exception as e:
            print(f"[API] Auth: web-form attempt failed: {e}")

    # 3. HTTP Basic fallback
    if not authed:
        session.auth = (user, password)
        print("[API] Auth: HTTP Basic fallback")

    # ── API version: probe v2 first (Airflow 3.x), then v1 ──────────────────
    api_prefix = None
    for prefix in ("/api/v2", "/api/v1"):
        probe = session.get(f"{base}{prefix}/connections", params={"limit": 1})
        if probe.status_code != 404:
            api_prefix = prefix
            print(f"[API] Detected API prefix: {prefix}  (probe status {probe.status_code})")
            break
    if api_prefix is None:
        sys.exit("[ERROR] Could not detect Airflow API version — neither /api/v2 nor /api/v1 responded")

    # ── Whoami ────────────────────────────────────────────────────────────────
    for me_path in (f"{api_prefix}/me", f"{api_prefix}/currentUser"):
        me = session.get(f"{base}{me_path}")
        if me.ok:
            mj = me.json()
            uname = mj.get("username") or mj.get("email") or "?"
            roles  = [r.get("name") for r in mj.get("roles", [])]
            print(f"[API] Logged in as: {uname}  roles: {roles}")
            break

    def get_all(endpoint: str, key: str) -> list[dict] | None:
        """Returns list of items, or None if access denied/not found."""
        items, offset, limit = [], 0, 100
        while True:
            r = session.get(f"{base}{api_prefix}{endpoint}", params={"limit": limit, "offset": offset})
            if r.status_code in (401, 403):
                print(f"  [SKIP] {r.status_code} {r.reason}")
                print(f"  Headers: {dict(r.headers)}")
                print(f"  Body:    {r.text[:500]}")
                return None
            if r.status_code == 404:
                print(f"  [SKIP] 404 — endpoint {endpoint} not found in {api_prefix}")
                return None
            r.raise_for_status()
            body = r.json()
            batch = body.get(key, [])
            items.extend(batch)
            total = body.get("total_entries", body.get("count", len(items)))
            if len(items) >= total:
                break
            offset += limit
        return items

    data = AirflowData()

    # ── Connections ───────────────────────────────────────────────────────────
    print("[API] Fetching connections...")
    # v2 uses 'connection_id', v1 also uses 'connection_id'
    raw_conns = get_all("/connections", "connections")
    for c in (raw_conns or []):
        data.connections.append({
            "conn_id":   c.get("connection_id", c.get("conn_id", "")),
            "conn_type": c.get("conn_type", ""),
            "host":      c.get("host", ""),
            "port":      c.get("port", ""),
            "schema":    c.get("schema", ""),
            "login":     c.get("login", ""),
            "password":  c.get("password", ""),
            "extra":     json.dumps(c.get("extra", "")) if c.get("extra") else "",
        })

    # ── Variables ─────────────────────────────────────────────────────────────
    print("[API] Fetching variables...")
    raw_vars = get_all("/variables", "variables")
    for v in (raw_vars or []):
        data.variables.append({
            "key":         v.get("key", ""),
            "value":       v.get("value", ""),
            "description": v.get("description", ""),
        })

    # ── Config / secrets ──────────────────────────────────────────────────────
    print("[API] Fetching config...")
    r = session.get(f"{base}{api_prefix}/config")
    if r.status_code in (401, 403):
        print(f"  [SKIP] config {r.status_code} — insufficient permissions")
    elif r.status_code == 404:
        print(f"  [SKIP] config 404 — not available in {api_prefix}")
    elif r.ok:
        cfg = r.json()
        for section in cfg.get("sections", []):
            sec_name = section.get("name", "")
            for opt in section.get("options", []):
                data.secrets.append({
                    "section": sec_name,
                    "key":     opt.get("key", ""),
                    "value":   opt.get("value", ""),
                    "source":  "airflow_config",
                })
    else:
        print(f"  [SKIP] config {r.status_code}: {r.text[:100]}")

    return data


# ─────────────────────────── LOCAL MODE ──────────────────────────────────────

def run_airflow_cli(args: list[str], env_extra: dict | None = None) -> str:
    import os
    env = os.environ.copy()
    if env_extra:
        env.update(env_extra)
    result = subprocess.run(
        ["airflow"] + args,
        capture_output=True, text=True, env=env,
    )
    if result.returncode != 0:
        raise RuntimeError(f"airflow {' '.join(args)} failed:\n{result.stderr}")
    return result.stdout


def fetch_local(airflow_home: str | None = None) -> AirflowData:
    import os
    env_extra = {}
    if airflow_home:
        env_extra["AIRFLOW_HOME"] = airflow_home

    data = AirflowData()

    print("[LOCAL] Fetching connections...")
    try:
        out = run_airflow_cli(["connections", "export", "/dev/stdout", "--file-format", "json"], env_extra)
        conns = extract_json(out)
        if isinstance(conns, list):
            items = conns
        else:
            items = list(conns.values())
        for c in items:
            data.connections.append({
                "conn_id":   c.get("conn_id", ""),
                "conn_type": c.get("conn_type", ""),
                "host":      c.get("host", ""),
                "port":      c.get("port", ""),
                "schema":    c.get("schema", ""),
                "login":     c.get("login", ""),
                "password":  c.get("password", ""),
                "extra":     json.dumps(c.get("extra", "")) if c.get("extra") else "",
            })
    except Exception as e:
        print(f"[LOCAL] connections failed: {e}")

    print("[LOCAL] Fetching variables...")
    try:
        out = run_airflow_cli(["variables", "export", "/dev/stdout"], env_extra)
        variables = extract_json(out)
        for k, v in variables.items():
            data.variables.append({"key": k, "value": str(v), "description": ""})
    except Exception as e:
        print(f"[LOCAL] variables failed: {e}")

    print("[LOCAL] Reading airflow.cfg for secrets/config...")
    try:
        home = airflow_home or os.environ.get("AIRFLOW_HOME", os.path.expanduser("~/airflow"))
        cfg_path = Path(home) / "airflow.cfg"
        if cfg_path.exists():
            import configparser
            cfg = configparser.ConfigParser()
            cfg.read(cfg_path)
            for section in cfg.sections():
                for key, value in cfg.items(section):
                    data.secrets.append({
                        "section": section,
                        "key":     key,
                        "value":   value,
                        "source":  str(cfg_path),
                    })
        else:
            print(f"[LOCAL] airflow.cfg not found at {cfg_path}")
    except Exception as e:
        print(f"[LOCAL] cfg read failed: {e}")

    return data


# ─────────────────────────── DOCKER MODE ─────────────────────────────────────

def fetch_docker(container: str, airflow_home: str = "/opt/airflow") -> AirflowData:
    import docker as docker_sdk

    client = docker_sdk.from_env()

    def exec_in_container(cmd: str) -> str:
        result = client.containers.get(container).exec_run(
            ["/bin/bash", "-c", cmd], demux=True
        )
        stdout, stderr = result.output
        if result.exit_code != 0:
            raise RuntimeError(f"cmd failed (exit {result.exit_code}): {(stderr or b'').decode()}")
        return (stdout or b"").decode()

    data = AirflowData()

    print(f"[DOCKER] Fetching connections from container '{container}'...")
    try:
        out = exec_in_container("airflow connections export /dev/stdout --file-format json 2>/dev/null")
        conns = extract_json(out)
        if isinstance(conns, list):
            items = conns
        else:
            items = list(conns.values())
        for c in items:
            data.connections.append({
                "conn_id":   c.get("conn_id", ""),
                "conn_type": c.get("conn_type", ""),
                "host":      c.get("host", ""),
                "port":      c.get("port", ""),
                "schema":    c.get("schema", ""),
                "login":     c.get("login", ""),
                "password":  c.get("password", ""),
                "extra":     json.dumps(c.get("extra", "")) if c.get("extra") else "",
            })
    except Exception as e:
        print(f"[DOCKER] connections failed: {e}")

    print(f"[DOCKER] Fetching variables from container '{container}'...")
    try:
        out = exec_in_container("airflow variables export /dev/stdout 2>/dev/null")
        variables = extract_json(out)
        for k, v in variables.items():
            data.variables.append({"key": k, "value": str(v), "description": ""})
    except Exception as e:
        print(f"[DOCKER] variables failed: {e}")

    print(f"[DOCKER] Reading airflow.cfg from container '{container}'...")
    try:
        cfg_text = exec_in_container(f"cat {airflow_home}/airflow.cfg 2>/dev/null || cat /root/airflow/airflow.cfg 2>/dev/null")
        import configparser, io
        cfg = configparser.ConfigParser()
        cfg.read_file(io.StringIO(cfg_text))
        for section in cfg.sections():
            for key, value in cfg.items(section):
                data.secrets.append({
                    "section": section,
                    "key":     key,
                    "value":   value,
                    "source":  f"docker:{container}:{airflow_home}/airflow.cfg",
                })
    except Exception as e:
        print(f"[DOCKER] cfg read failed: {e}")

    print(f"[DOCKER] Fetching env vars from container '{container}'...")
    try:
        env_out = exec_in_container("env")
        for line in env_out.splitlines():
            if "=" in line:
                k, _, v = line.partition("=")
                if any(kw in k.upper() for kw in ("SECRET", "PASSWORD", "TOKEN", "KEY", "AIRFLOW__", "FERNET")):
                    data.secrets.append({
                        "section": "env",
                        "key":     k,
                        "value":   v,
                        "source":  f"docker:{container}:env",
                    })
    except Exception as e:
        print(f"[DOCKER] env vars failed: {e}")

    return data


# ─────────────────────────── EXCEL EXPORT ────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ALT_FILL    = PatternFill("solid", fgColor="D6E4F0")


def _write_sheet(ws, headers: list[str], rows: list[dict]):
    ws.append(headers)
    for cell in ws[1]:
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for i, row in enumerate(rows, start=2):
        ws.append([str(row.get(h, "")) for h in headers])
        if i % 2 == 0:
            for cell in ws[i]:
                cell.fill = ALT_FILL

    for col_idx, _ in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            (len(str(ws.cell(r, col_idx).value or "")) for r in range(1, ws.max_row + 1)),
            default=10,
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)


def export_excel(data: AirflowData, output: str):
    wb = openpyxl.Workbook()

    # Connections
    ws_conn = wb.active
    ws_conn.title = "Connections"
    _write_sheet(ws_conn,
        ["conn_id", "conn_type", "host", "port", "schema", "login", "password", "extra"],
        data.connections,
    )

    # Variables
    ws_vars = wb.create_sheet("Variables")
    _write_sheet(ws_vars,
        ["key", "value", "description"],
        data.variables,
    )

    # Secrets / Config
    ws_sec = wb.create_sheet("Secrets_Config")
    _write_sheet(ws_sec,
        ["section", "key", "value", "source"],
        data.secrets,
    )

    # Summary
    ws_sum = wb.create_sheet("Summary")
    ws_sum.append(["Sheet", "Count"])
    ws_sum["A1"].font = HEADER_FONT
    ws_sum["A1"].fill = HEADER_FILL
    ws_sum["B1"].font = HEADER_FONT
    ws_sum["B1"].fill = HEADER_FILL
    ws_sum.append(["Connections",   len(data.connections)])
    ws_sum.append(["Variables",     len(data.variables)])
    ws_sum.append(["Secrets/Config", len(data.secrets)])
    ws_sum.column_dimensions["A"].width = 20
    ws_sum.column_dimensions["B"].width = 10

    wb.save(output)
    print(f"\nSaved: {output}")
    print(f"  Connections:    {len(data.connections)}")
    print(f"  Variables:      {len(data.variables)}")
    print(f"  Secrets/Config: {len(data.secrets)}")


# ─────────────────────────── CLI ─────────────────────────────────────────────

def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(
        description="Extract Airflow connections, variables & secrets to Excel"
    )
    p.add_argument("--mode", required=True, choices=["api", "local", "docker"],
                   help="Extraction mode")
    p.add_argument("--output", default="airflow_export.xlsx",
                   help="Output .xlsx file (default: airflow_export.xlsx)")

    # API args
    ag = p.add_argument_group("API mode")
    ag.add_argument("--url",      help="Airflow base URL, e.g. http://localhost:8080")
    ag.add_argument("--user",     default="admin", help="Airflow username")
    ag.add_argument("--password", default="admin", help="Airflow password")
    ag.add_argument("--cookie",   help="Raw browser cookie string, e.g. 'session=abc123' — bypasses login")

    # Local args
    lg = p.add_argument_group("Local mode")
    lg.add_argument("--airflow-home", help="AIRFLOW_HOME path (default: ~/airflow)")

    # Docker args
    dg = p.add_argument_group("Docker mode")
    dg.add_argument("--container",    help="Docker container name or ID")
    dg.add_argument("--docker-home",  default="/opt/airflow",
                    help="AIRFLOW_HOME inside container (default: /opt/airflow)")

    return p


def main():
    parser = build_parser()
    args = parser.parse_args()

    if args.mode == "api":
        if not args.url:
            parser.error("--url is required for api mode")
        data = fetch_api(args.url, args.user, args.password, cookie=args.cookie)

    elif args.mode == "local":
        data = fetch_local(args.airflow_home)

    elif args.mode == "docker":
        if not args.container:
            parser.error("--container is required for docker mode")
        data = fetch_docker(args.container, args.docker_home)

    else:
        parser.error(f"Unknown mode: {args.mode}")

    export_excel(data, args.output)


if __name__ == "__main__":
    main()
